# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     wechat_collector.py
   Description :   从Excel读取公众号配置，按配置的小时定时执行采集
   Author :        Xiaoxing (Modified)
   date：          2026/05/15
-------------------------------------------------
"""

import requests
from loguru import logger
from datetime import datetime, timezone, timedelta
import re
import time
import json
import openpyxl
import sys

# ==================== 全局配置 ====================
MAX_RETRY_COUNT = 3          # 最大重试次数
TIMEOUT = 10                 # 请求超时时间（秒）
MAX_PAGE = 10                # 历史文章最大翻页数
COMMENT_MAX_PAGE = 10        # 评论最多翻页数

# 接口密钥（需要用户自行填写）
jzlkey = "你的jzlkey"
verify_code = ""

# 时区设置（北京时间）
BJ_TZ = timezone(timedelta(hours=8))

# ==================== 工具函数 ====================
def parse_hours(hours_str):
    """
    解析采集时间点字符串，返回整数小时列表
    支持中英文逗号分隔，例如 "11,22" 或 "11，22"
    """
    if not hours_str:
        return []
    # 替换中文逗号为英文逗号
    hours_str = hours_str.replace('，', ',')
    parts = hours_str.split(',')
    hours = []
    for p in parts:
        p = p.strip()
        if p.isdigit():
            h = int(p)
            if 0 <= h <= 23:
                hours.append(h)
    return hours

def calculate_time_range(range_type_text):
    """
    根据时间范围类型文本计算起止时间字符串（北京时间）
    :param range_type_text: "当天" 或 "过去24小时"
    :return: (start_time_str, end_time_str) 格式 "%Y-%m-%d %H:%M:%S"
    """
    now = datetime.now(BJ_TZ)
    if range_type_text == "当天":
        # 当天 00:00:00
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif range_type_text == "过去24小时":
        start = now - timedelta(hours=24)
        end = now
    else:
        raise ValueError(f"不支持的时间范围类型: {range_type_text}，请使用'当天'或'过去24小时'")
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")

def parse_bool_flag(value):
    """
    将 Excel 中的值（"是"/"否" 或 1/0）转换为布尔值
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    if isinstance(value, str):
        return value.strip() in ("是", "1", "yes", "true", "True")
    return False

# ==================== API 调用函数 ====================
def get_mp_history(nickname: str, biz: str, start_time: str, end_time: str):
    """获取公众号历史文章列表"""
    api_url = "https://www.dajiala.com/fbmain/monitor/v3/post_history"

    post_data = {
        "biz": biz if biz else "",
        "url": "",
        "name": nickname,
        "page": 1,
        "key": jzlkey,
        "verifycode": verify_code
    }

    start_ts = int(datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S").replace(tzinfo=BJ_TZ).timestamp())
    end_ts = int(datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S").replace(tzinfo=BJ_TZ).timestamp())

    is_over = False
    articles_list = []
    for page in range(1, MAX_PAGE + 1):
        if is_over:
            logger.info("已达到结束时间，停止翻页")
            break
        post_data["page"] = page
        for retry in range(MAX_RETRY_COUNT):
            try:
                response = requests.post(api_url, json=post_data, timeout=TIMEOUT).json()
                if response.get("code") == 0:
                    logger.info(f"第{page}页数据获取成功")
                    data = response.get("data", [])
                    for article in data:
                        article_url = article.get("url", "")
                        pattern = re.compile(r'https://mp.weixin.qq.com/s/([\S]{22})')
                        match = pattern.match(article_url)
                        if match:
                            article_id = match.group(1)
                        else:
                            logger.warning(f"文章链接格式不正确，无法提取article_id: {article_url}")
                            continue
                        post_time = article.get("post_time", 0)
                        if post_time < start_ts:
                            logger.info("文章发布时间早于开始时间，停止翻页")
                            is_over = True
                            break
                        if post_time > end_ts:
                            logger.info("文章发布时间晚于结束时间，跳过不收集")
                            continue
                        post_time_str = article.get("post_time_str", "")
                        cover_url = article.get("cover_url", "")
                        is_deleted = article.get("is_deleted", "0")
                        msg_status = article.get("msg_status", 2)
                        # 过滤已删除/违规文章
                        if msg_status == 2 and is_deleted == "0":
                            articles_list.append({
                                "original_url": article_url,
                                "article_id": article_id,
                                "title": article.get("title", ""),
                                "post_time": post_time,
                                "post_time_str": post_time_str,
                                "cover_url": cover_url
                            })
                    break  # 成功，跳出重试循环
                elif response.get("code") == 105:
                    logger.warning("无法搜索该公众号，请检查公众号名称是否正确")
                    is_over = True
                    break
                else:
                    logger.warning(f"API返回错误码: {response.get('code')}, 信息: {response.get('msg')}")
                    if retry == MAX_RETRY_COUNT - 1:
                        logger.error("达到最大重试次数，停止翻页")
                        is_over = True
            except Exception as e:
                logger.error(f"请求异常: {e}")
                if retry == MAX_RETRY_COUNT - 1:
                    logger.error("达到最大重试次数，停止尝试")
                    is_over = True
    return articles_list

def get_article_info(article_url: str):
    """获取文章详情（正文、作者等）"""
    api_url = "https://www.dajiala.com/fbmain/monitor/v3/article_detail"
    post_data = {
        "url": article_url,
        "key": jzlkey,
        "mode": 2,
        "verifycode": verify_code
    }

    for retry in range(MAX_RETRY_COUNT):
        try:
            response = requests.get(api_url, params=post_data, timeout=TIMEOUT).json()
            if response.get("code") == 0:
                logger.info("文章详情获取成功")
                return {
                    "ghid": response.get("user_name", ""),
                    "biz": response.get("biz", ""),
                    "mp_nickname": response.get("nick_name", ""),
                    "content_text": response.get("content", ""),
                    "html_content": response.get("content_multi_text", ""),
                    "author": response.get("author", "")
                }
            elif response.get("code") == 101:
                logger.warning("文章无法访问，可能已删除或违规")
                return {}
        except Exception as e:
            logger.error(f"请求异常: {e}")
            if retry == MAX_RETRY_COUNT - 1:
                logger.error("已达到最大重试次数，停止尝试")
                return {}
    return {}

def get_read_zan_data(article_url: str):
    """获取阅读、点赞、在看、转发、评论数"""
    api_url = "https://www.dajiala.com/fbmain/monitor/v3/read_zan_pro"
    post_data = {
        "url": article_url,
        "key": jzlkey,
        "mode": 1,
        "verifycode": verify_code
    }

    for retry in range(MAX_RETRY_COUNT):
        try:
            response = requests.post(api_url, json=post_data, timeout=TIMEOUT).json()
            if response.get("code") == 0:
                logger.info("文章互动数据获取成功")
                data = response.get("data", {})
                return {
                    "description": "互动指标对象",
                    "read_count": data.get("read", 0),
                    "like_count": data.get("zan", 0),
                    "wow_count": data.get("looking", 0),
                    "forward_count": data.get("share_num", 0),
                    "comment_count": data.get("comment_count", 0)
                }
            elif response.get("code") == 101:
                logger.warning("文章无法访问，可能已删除或违规")
                return {
                    "description": "互动指标对象",
                    "read_count": 0,
                    "like_count": 0,
                    "wow_count": 0,
                    "forward_count": 0,
                    "comment_count": 0
                }
        except Exception as e:
            logger.error(f"请求异常: {e}")
            if retry == MAX_RETRY_COUNT - 1:
                logger.error("已达到最大重试次数，停止尝试")
                return {}
    return {}

def get_article_comments(article_url: str):
    """获取文章评论（一页最多100条，最多翻COMMENT_MAX_PAGE页）"""
    api_url = "https://www.dajiala.com/fbmain/monitor/v3/article_comment2"
    post_data = {
        "url": article_url,
        "buffer": "",
        "key": jzlkey,
        "verifycode": verify_code
    }
    comments_list = []
    is_over = False

    for page in range(1, COMMENT_MAX_PAGE + 1):
        if is_over:
            break
        for retry in range(MAX_RETRY_COUNT):
            try:
                response = requests.post(api_url, data=post_data)
                response.raise_for_status()
                result = response.json()
                if result.get("code") == 0:
                    buffer = result.get("buffer", "")
                    continue_flag = result.get("continue_flag", False)
                    logger.info(f"第 {page} 页评论获取成功，本页评论数：{len(result.get('data', []))}")
                    for item in result.get("data", []):
                        comments_list.append({
                            "description": "评论列表对象",
                            "comment_id": item.get("content_id", ""),
                            "user_nickname": item.get("nick_name", ""),
                            "comment_text": item.get("content", ""),
                            "like_count": item.get("like_num", 0),
                            "publish_time": item.get("create_time_stamp", 0)
                        })
                    if not continue_flag or len(result.get("data", [])) < 100:
                        logger.info("评论获取完毕，没有更多评论")
                        is_over = True
                    else:
                        post_data["buffer"] = buffer
                    break  # 成功，跳出重试循环
                elif result.get("code") == 101:
                    logger.warning("文章无法访问，可能已删除或违规，无法获取评论")
                    is_over = True
                    break
                else:
                    logger.error(f"获取评论失败，code: {result.get('code')}")
            except Exception as e:
                logger.error(f"请求异常: {e}")
                if retry == MAX_RETRY_COUNT - 1:
                    logger.error("已达到最大重试次数，停止尝试")
                    is_over = True
    return comments_list

# ==================== 主处理函数 ====================
def process_account(row_data, current_hour):
    """
    处理单个公众号的采集任务（带时间点判断）
    row_data: dict 包含 name, category, time_range_type, need_interact, need_comment, hours_list
    current_hour: 当前小时（0-23）
    返回采集到的文章列表（可能为空）
    """
    nickname = row_data["name"]
    category = row_data["category"]
    time_range_text = row_data["time_range_type"]
    need_interact = row_data["need_interact"]   # bool
    need_comment = row_data["need_comment"]     # bool
    hours_list = row_data.get("hours_list", [])

    # 时间点判断：如果指定了小时列表且当前小时不在列表中，则跳过
    if hours_list and current_hour not in hours_list:
        logger.info(f"公众号 {nickname} 当前小时 {current_hour} 不在采集时间点 {hours_list} 内，跳过")
        return []

    # 计算时间范围
    try:
        start_time, end_time = calculate_time_range(time_range_text)
    except ValueError as e:
        logger.error(f"公众号 {nickname} 时间范围类型错误: {e}")
        return []

    logger.info(f"开始处理公众号: {nickname}, 时间范围: {start_time} 至 {end_time}")

    # 第一步：获取历史文章列表（biz参数先传空，第一次用昵称查询）
    articles = get_mp_history(nickname, "", start_time, end_time)
    if not articles:
        logger.warning(f"公众号 {nickname} 未获取到文章")
        return []
    logger.info(f"共获取到 {len(articles)} 篇文章")

    results = []
    for article in articles:
        original_url = article["original_url"]
        logger.info(f"正在处理文章: {article['title']}")

        # 获取文章详情（始终获取，因为需要作者、正文等）
        article_info = get_article_info(original_url)
        if not article_info:
            logger.warning(f"文章详情获取失败，跳过该文章")
            continue

        # 获取互动数据（根据配置）
        if need_interact:
            interact_data = get_read_zan_data(original_url)
        else:
            interact_data = None
            logger.info("根据配置不获取互动数据")

        # 获取评论（根据配置且当评论数>0时）
        if need_comment and interact_data.get("comment_count", 0) > 0:
            comments = get_article_comments(original_url)
        else:
            comments = []
            if need_comment and interact_data.get("comment_count", 0) == 0:
                logger.info("文章无评论，不获取评论列表")
            elif not need_comment:
                logger.info("根据配置不获取评论数据")

        # 组装数据
        temp_data = {
            "article_id": article.get("article_id", ""),
            "source_account": article_info.get("mp_nickname", nickname),
            "category": category,
            "title": article.get("title", ""),
            "author": article_info.get("author", ""),
            "publish_time": article.get("post_time", 0),
            "fetch_time": int(time.time()),
            "original_url": original_url,
            "content": {
                "html": article_info.get("html_content", ""),
                "text": article_info.get("content_text", "")
            },
            "engagement_data": interact_data,
            "comments": comments
        }
        results.append(temp_data)

    # 保存该公众号的结果到独立的JSON文件（文件名包含日期小时，避免覆盖）
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", nickname)
    time_suffix = datetime.now(BJ_TZ).strftime("%Y%m%d_%H")
    out_file = f"{safe_name}_{time_suffix}.json"

    # 保存结果  如果是保存数据库 请修改对应的代码
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=4)
    logger.info(f"公众号 {nickname} 结果已保存至 {out_file}")
    return results

def load_config_from_excel(excel_path):
    """
    使用 openpyxl 读取 Excel 配置文件
    期望列名（完全匹配）：
        公众号名称, 类别(category), 时间范围类型, 是否获取互动数据, 是否获取评论数据, 采集时间点
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    # 获取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    # 期望的列名映射
    required = ["公众号名称", "类别(category)", "时间范围类型", "是否获取互动数据", "是否获取评论数据", "采集时间点"]
    col_index = {}
    for req in required:
        if req not in headers:
            raise ValueError(f"Excel缺少必需列: {req}")
        col_index[req] = headers.index(req)
    # 读取数据行
    configs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        name = row[col_index["公众号名称"]]
        category = row[col_index["类别(category)"]]
        time_type = row[col_index["时间范围类型"]]
        need_interact_val = row[col_index["是否获取互动数据"]]
        need_comment_val = row[col_index["是否获取评论数据"]]
        hours_str = row[col_index["采集时间点"]]

        if not name:
            logger.warning("公众号名称为空，跳过该行")
            continue

        # 解析时间点列表
        hours_list = parse_hours(str(hours_str) if hours_str else "")

        configs.append({
            "name": str(name).strip(),
            "category": str(category).strip() if category else "",
            "time_range_type": str(time_type).strip() if time_type else "",
            "need_interact": parse_bool_flag(need_interact_val),
            "need_comment": parse_bool_flag(need_comment_val),
            "hours_list": hours_list
        })
    return configs

# ==================== 主入口 ====================
if __name__ == "__main__":
    # 请在此处指定 Excel 文件路径
    EXCEL_PATH = "wechat_config.xlsx"   # 修改为实际路径

    # 获取当前小时（北京时间）
    now = datetime.now(BJ_TZ)
    current_hour = now.hour
    logger.info(f"当前时间: {now.strftime('%Y-%m-%d %H:%M:%S')}, 当前小时: {current_hour}")

    try:
        config_list = load_config_from_excel(EXCEL_PATH)
        logger.info(f"已读取 {len(config_list)} 个公众号配置 {config_list}")
    except Exception as e:
        logger.error(f"读取Excel失败: {e}")
        sys.exit(1)

    if not config_list:
        logger.warning("Excel中没有有效配置，程序退出")
        sys.exit(0)
    
    all_results = []
    for cfg in config_list:
        result = process_account(cfg, current_hour)
        if result:
            all_results.extend(result)

    logger.info(f"全部处理完成，本次共采集 {len(all_results)} 篇文章")